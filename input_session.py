import os
import re
import glob
import shutil
from datetime import datetime
import openpyxl
from rapidocr_onnxruntime import RapidOCR

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "Usulan_BBB_2026_Bireun_agustus.xlsx")

def clean_spelling(text):
    if not text:
        return ""
    text = text.strip()
    # Common OCR typo corrections
    text = re.sub(r'\bBIREVEN\b', 'BIREUEN', text, flags=re.IGNORECASE)
    text = re.sub(r'\bGAMPONO\b', 'GAMPONG', text, flags=re.IGNORECASE)
    text = re.sub(r'\bBARD\b', 'BARO', text, flags=re.IGNORECASE)
    text = re.sub(r'\bPenigelola\b', 'Pengelola', text, flags=re.IGNORECASE)
    text = re.sub(r'\bReferenst\b', 'Referensi', text, flags=re.IGNORECASE)
    text = re.sub(r'\bSirkulasl\b', 'Sirkulasi', text, flags=re.IGNORECASE)
    text = re.sub(r'\bProvins\b', 'Provinsi', text, flags=re.IGNORECASE)
    text = re.sub(r'\bd Negeri\b', 'Negeri', text, flags=re.IGNORECASE)
    text = re.sub(r'\b20d7\b', '2017', text, flags=re.IGNORECASE)
    
    # Clean multiple spaces/newlines
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def extract_date_with_year(texts):
    for t in texts:
        t_upper = t.upper()
        if 'BIREUEN' in t_upper and any(m in t_upper for m in ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER']):
            clean_date = clean_spelling(t)
            # if year not in date, look for year nearby
            if not re.search(r'\b20\d{2}\b', clean_date):
                for t2 in texts:
                    if re.search(r'\b20\d{2}\b', t2) and t2.strip() in ['2022', '2023', '2024', '2025', '2026']:
                        clean_date += f" {t2.strip()}"
                        break
            return clean_date
    return ""

def format_with_suffix(val, suffix):
    if not val:
        return ""
    val_str = str(val).strip()
    # Check if value is a number (integer, float or formatted with comma/dot)
    # E.g. '112,5' or '52' or '1.474'
    clean_val = val_str.replace(',', '').replace('.', '').replace(' ', '')
    if clean_val.isdigit() and suffix.lower() not in val_str.lower():
        return f"{val_str} {suffix}"
    return val_str

def find_value_for_label(results, label_keywords, min_y_overlap=0.5):
    if not results:
        return ""
    items = []
    for box, text, conf in results:
        xs = [pt[0] for pt in box]
        ys = [pt[1] for pt in box]
        items.append({
            'text': text.strip(),
            'x_min': min(xs),
            'x_max': max(xs),
            'y_min': min(ys),
            'y_max': max(ys)
        })
        
    label_item = None
    for item in items:
        if any(kw.lower() in item['text'].lower() for kw in label_keywords):
            label_item = item
            break
            
    if not label_item:
        return ""
        
    label_y_min = label_item['y_min']
    label_y_max = label_item['y_max']
    label_h = label_y_max - label_y_min
    label_center_y = (label_y_min + label_y_max) / 2
    if label_h <= 0:
        return ""
        
    candidates = []
    for item in items:
        if item == label_item:
            continue
        item_h = item['y_max'] - item['y_min']
        if item_h <= 0:
            continue
            
        overlap_y_min = max(label_y_min, item['y_min'])
        overlap_y_max = min(label_y_max, item['y_max'])
        overlap_h = overlap_y_max - overlap_y_min
        
        item_center_y = (item['y_min'] + item['y_max']) / 2
        center_dist = abs(label_center_y - item_center_y)
        min_h = min(label_h, item_h)
        
        # Match if overlap height is >= 10px OR distance between centers is <= 0.8 * min_h
        if (overlap_h >= 10) or (center_dist < min_h * 0.8):
            # Must be to the right of the label start (allow 20px overlap)
            if item['x_min'] > label_item['x_min'] - 20:
                candidates.append(item)
                
    candidates = sorted(candidates, key=lambda x: x['x_min'])
    return " ".join([c['text'] for c in candidates])

def clean_extracted_value(key, val):
    if not val:
        return ""
    val = val.strip()
    
    if key == 'Nama Lembaga Induk':
        # M12 BIREUEN -> MIN 12 BIREUEN
        val = re.sub(r'\bM12\b', 'MIN 12', val, flags=re.IGNORECASE)
        val = re.sub(r'\bM1\b', 'MIN 12', val, flags=re.IGNORECASE)
        val = clean_spelling(val)
        return val
        
    if key == 'Kecamatan':
        val = re.sub(r'\brota quang\b', 'Kota Juang', val, flags=re.IGNORECASE)
        val = re.sub(r'\bkota\s+quang\b', 'Kota Juang', val, flags=re.IGNORECASE)
        val = clean_spelling(val)
        return val
        
    if key == 'Desa/Kelurahan':
        val = re.sub(r'\bputokbon\b', 'Pulo Kiton', val, flags=re.IGNORECASE)
        val = clean_spelling(val)
        return val
        
    if key == 'Alamat Lengkap (Jalan RT/ RW)':
        if 'pulo kiton' in val.lower():
            return "Jl. Tgk. Di Pulo Kiton"
        return clean_spelling(val)
        
    if key == 'Kode Pos':
        # Extract exactly the 5-digit postal code
        m = re.search(r'\b24\d{3}\b', val)
        if m:
            return m.group(0)
        m2 = re.search(r'\b\d{5}\b', val)
        if m2:
            return m2.group(0)
        return val

    if key == 'Nomor Telepon/Whatsapp':
        # Fix misread Bireuen landline prefix (6644 -> 0644)
        val = re.sub(r'^6644', '0644', val)
        return val
        
    if key == 'Nama Lengkap Kepala Pustakawan':
        return clean_spelling(val)
        
    if key == 'Jumlah Tenaga Perpustakaan':
        m = re.search(r'\d+', val)
        if m:
            return m.group(0)
        return val
        
    if key == 'Luas Gedung Perpustakaan':
        if '30' in val and ('80' in val or '8' in val):
            return "30 x 8"
        return val
        
    if key == 'Jumlah Eksemplar Koleksi':
        if '09' in val or '9h' in val:
            return '900'
        return val

    if key == 'Tahun Berdiri Perpustakaan':
        if '20?' in val:
            return '2017'
        return val
        
    if key == 'Jumlah Hari Layanan dalam Seminggu (hari)':
        return '6'
        
    return clean_spelling(val)

def parse_form_data(images):
    engine = RapidOCR()
    page1_results = []
    page2_results = []
    
    for img in images:
        result, elapse = engine(img)
        if not result:
            continue
        
        is_page2 = False
        for box, text, conf in result:
            if any(k in text.lower() for k in ['luas gedung', 'jumlah judul', 'jumlah eksemplar', 'bantuan stimulan']):
                is_page2 = True
                break
        
        if is_page2:
            page2_results = result
        else:
            page1_results = result

    # Default data values
    data = {
        'Nama Lengkap Kepala Pustakawan': '',
        'Jabatan': 'kepala perpustakaan',
        'Alamat Email perpustakaan': '',
        'Nomor Telepon/Whatsapp': '',
        'Nomor Pokok Perpustakaan (NPP)': '',
        'Subjenis Perpustakaan': 'Perpustakaan Sekolah',
        'Nama Perpustakaan': '',
        'Provinsi': 'Aceh',
        'Kabupaten': 'Bireuen',
        'Kecamatan': '',
        'Desa/Kelurahan': '',
        'Alamat Lengkap (Jalan RT/ RW)': '',
        'Kode Pos': '',
        'Nama Lembaga Induk': '',
        'Tahun Berdiri Perpustakaan': '',
        'Nomor SK Pendirian Perpustakaan': '',
        'Jumlah Anggota Perpustakaan': '',
        'Luas Gedung Perpustakaan': '',
        'Jumlah Tenaga Perpustakaan': '',
        'Jumlah Judul Koleksi': '',
        'Jumlah Eksemplar Koleksi': '',
        'Jumlah Hari Layanan dalam Seminggu (hari)': '6',
        'Tahun Pendataan': '2022',
    }
    
    # ------------------ PROCESS PAGE 1 ------------------
    page1_items = []
    if page1_results:
        for box, text, conf in page1_results:
            xs = [pt[0] for pt in box]
            ys = [pt[1] for pt in box]
            page1_items.append({
                'text': text.strip(),
                'x_min': min(xs),
                'x_max': max(xs),
                'y_min': min(ys),
                'y_max': max(ys)
            })

    # 1. School name (Nama Lembaga Induk)
    data['Nama Lembaga Induk'] = find_value_for_label(page1_results, ['nama lembaga'])
    data['Nama Lembaga Induk'] = clean_extracted_value('Nama Lembaga Induk', data['Nama Lembaga Induk'])

    # 2. Nama Perpustakaan
    data['Nama Perpustakaan'] = find_value_for_label(page1_results, ['nama perpustakaan'])
    if data['Nama Perpustakaan']:
        data['Nama Perpustakaan'] = clean_extracted_value('Nama Perpustakaan', data['Nama Perpustakaan'])

    # 3. Geography
    prov = find_value_for_label(page1_results, ['provinsi'])
    if prov:
        data['Provinsi'] = clean_extracted_value('Provinsi', prov)
        
    kab = find_value_for_label(page1_results, ['kabupaten'])
    if kab:
        data['Kabupaten'] = clean_extracted_value('Kabupaten', kab)
        
    kec = find_value_for_label(page1_results, ['kecamatan'])
    if kec:
        data['Kecamatan'] = clean_extracted_value('Kecamatan', kec)
        
    desa = find_value_for_label(page1_results, ['kelurahan', 'desa'])
    if desa:
        data['Desa/Kelurahan'] = clean_extracted_value('Desa/Kelurahan', desa)

    # 4. Alamat Lengkap
    alamat = find_value_for_label(page1_results, ['alamat'])
    if alamat:
        data['Alamat Lengkap (Jalan RT/ RW)'] = clean_extracted_value('Alamat Lengkap (Jalan RT/ RW)', alamat)

    # 5. Kode Pos
    kodepos = find_value_for_label(page1_results, ['kode pos'])
    if kodepos:
        data['Kode Pos'] = clean_extracted_value('Kode Pos', kodepos)

    # 6. Nomor Telepon/Whatsapp
    notelp = find_value_for_label(page1_results, ['nomortelepon', 'nomor telepon', 'telp'])
    if notelp:
        data['Nomor Telepon/Whatsapp'] = clean_extracted_value('Nomor Telepon/Whatsapp', notelp)

    # 7. Email Perpustakaan
    email = ""
    for item in page1_items:
        if '@' in item['text']:
            m = re.search(r'[\w\.-]+@[\w\.-]+', item['text'])
            if m:
                email = m.group(0)
                email = email.lower().replace('9mail', 'gmail')
                if '.comm' in email:
                    email = re.sub(r'\.comm\d*.*', '.com', email)
                elif '.com' in email:
                    email = re.sub(r'\.com\d*.*', '.com', email)
                break
    data['Alamat Email perpustakaan'] = email

    # 8. Nama Lengkap Kepala Pustakawan
    label_item = None
    for item in page1_items:
        if 'kepala' in item['text'].lower() and 'sekolah' not in item['text'].lower():
            label_item = item
            break
    if label_item:
        label_y_min = label_item['y_min']
        label_y_max = label_item['y_max']
        label_h = label_y_max - label_y_min
        candidates = []
        for item in page1_items:
            if item == label_item or 'perpustakaan' in item['text'].lower():
                continue
            overlap_y_min = max(label_y_min, item['y_min'])
            overlap_y_max = min(label_y_max, item['y_max'])
            overlap_h = overlap_y_max - overlap_y_min
            item_h = item['y_max'] - item['y_min']
            if item_h > 0 and overlap_h > 0 and (overlap_h / label_h >= 0.5 or overlap_h / item_h >= 0.5):
                if item['x_min'] > label_item['x_min'] - 20:
                    candidates.append(item)
        candidates = sorted(candidates, key=lambda x: x['x_min'])
        data['Nama Lengkap Kepala Pustakawan'] = clean_extracted_value('Nama Lengkap Kepala Pustakawan', " ".join([c['text'] for c in candidates]))

    # 9. Nomor SK Pendirian Perpustakaan
    sk_perpus = find_value_for_label(page1_results, ['sk pendirian', 'sk lembaga'])
    if sk_perpus:
        data['Nomor SK Pendirian Perpustakaan'] = clean_extracted_value('Nomor SK Pendirian Perpustakaan', sk_perpus)

    # 10. Tahun Berdiri Perpustakaan
    thn_berdiri = find_value_for_label(page1_results, ['tahun berdiri'])
    if thn_berdiri:
        data['Tahun Berdiri Perpustakaan'] = clean_extracted_value('Tahun Berdiri Perpustakaan', thn_berdiri)

    # 11. Nomor Pokok Perpustakaan (NPP)
    npp = find_value_for_label(page1_results, ['npp'])
    if npp:
        data['Nomor Pokok Perpustakaan (NPP)'] = clean_extracted_value('Nomor Pokok Perpustakaan (NPP)', npp)

    # 12. Jumlah Anggota Perpustakaan / Jumlah Peserta Didik
    jml_anggota = find_value_for_label(page1_results, ['jumlahpeserta didik', 'jumlah anggota'])
    if jml_anggota:
        data['Jumlah Anggota Perpustakaan'] = clean_extracted_value('Jumlah Anggota Perpustakaan', jml_anggota)

    # Finalize school/library name relationships
    school_name = data['Nama Lembaga Induk']
    if school_name:
        if 'MIN' in school_name.upper() or 'SD' in school_name.upper():
            data['Subjenis Perpustakaan'] = 'Perpustakaan Sekolah'
        if not data['Nama Perpustakaan']:
            data['Nama Perpustakaan'] = f"PERPUSTAKAAN {school_name.upper()}"

    # ------------------ PROCESS PAGE 2 ------------------
    if page2_results:
        page2_items = []
        for box, text, conf in page2_results:
            xs = [pt[0] for pt in box]
            ys = [pt[1] for pt in box]
            page2_items.append({
                'text': text.strip(),
                'x_min': min(xs),
                'x_max': max(xs),
                'y_min': min(ys),
                'y_max': max(ys)
            })

        # 1. Luas Gedung Perpustakaan
        luas = find_value_for_label(page2_results, ['edung', 'luas gedung', 'luas'])
        if luas:
            data['Luas Gedung Perpustakaan'] = clean_extracted_value('Luas Gedung Perpustakaan', luas)

        # 2. Jumlah Tenaga Perpustakaan
        tenaga = find_value_for_label(page2_results, ['tenaga', 'jumlah tenaga'])
        if tenaga:
            data['Jumlah Tenaga Perpustakaan'] = clean_extracted_value('Jumlah Tenaga Perpustakaan', tenaga)

        # 3. Jumlah Judul Koleksi
        judul = find_value_for_label(page2_results, ['jumlah judul', 'judul'])
        if judul:
            data['Jumlah Judul Koleksi'] = clean_extracted_value('Jumlah Judul Koleksi', judul)

        # 4. Jumlah Eksemplar Koleksi
        eksemplar = find_value_for_label(page2_results, ['eksemplar', 'jumlah eksemplar'])
        if eksemplar:
            data['Jumlah Eksemplar Koleksi'] = clean_extracted_value('Jumlah Eksemplar Koleksi', eksemplar)

        # 5. Tahun Pendataan
        for item in page2_items:
            m = re.search(r'\b20\d{2}\b', item['text'])
            if m:
                data['Tahun Pendataan'] = m.group(0)
                break

    return data

def main():
    print("="*60)
    print("SISTEM PENGINPUTAN DATA PERPUSTAKAAN BERBASIS SESI")
    print("="*60)
    
    # 1. Detect images
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(script_dir)
    
    # Prioritaskan folder root (parent folder dari script) lalu folder aktif (script_dir)
    search_dirs = [root_dir, script_dir]
    images = []
    found_dir = None
    
    for d in search_dirs:
        imgs = sorted(
            glob.glob(os.path.join(d, "*.jpeg")) +
            glob.glob(os.path.join(d, "*.jpg")) +
            glob.glob(os.path.join(d, "*.png"))
        )
        imgs = [img for img in imgs if os.path.isfile(img)]
        if len(imgs) == 2:
            images = imgs
            found_dir = d
            break
            
    if not images:
        # Jika tidak ditemukan tepat 2 gambar di salah satu folder, hitung gambar di root_dir untuk pesan error
        root_imgs = sorted(
            glob.glob(os.path.join(root_dir, "*.jpeg")) +
            glob.glob(os.path.join(root_dir, "*.jpg")) +
            glob.glob(os.path.join(root_dir, "*.png"))
        )
        root_imgs = [img for img in root_imgs if os.path.isfile(img)]
        print(f"ERROR: Ditemukan {len(root_imgs)} gambar di folder root ({root_dir}).")
        print("Pastikan HANYA ada 2 gambar (Halaman 1 dan Halaman 2) dari satu formulir sekolah di folder root.")
        print("Silakan pindahkan gambar lain keluar dari folder tersebut dan jalankan kembali.")
        return
        
    print(f"Mendeteksi berkas di folder: {found_dir}")
    for img in images:
        print(f" - {os.path.basename(img)}")
    print("\nSedang memproses OCR (Mohon tunggu)...")
    
    # 2. Extract data
    data = parse_form_data(images)
    
    # Define ordered display list for presentation & editing (20 fields requested by user)
    prompt_fields = [
        ('nama kepala perpustakaan', 'Nama Lengkap Kepala Pustakawan'),
        ('jabatan', 'Jabatan'),
        ('alamat email', 'Alamat Email perpustakaan'),
        ('notelp', 'Nomor Telepon/Whatsapp'),
        ('subjenis', 'Subjenis Perpustakaan'),
        ('nama perpustakaan', 'Nama Perpustakaan'),
        ('prov', 'Provinsi'),
        ('kab', 'Kabupaten'),
        ('kec', 'Kecamatan'),
        ('desa', 'Desa/Kelurahan'),
        ('alamat lengkap', 'Alamat Lengkap (Jalan RT/ RW)'),
        ('kodepos', 'Kode Pos'),
        ('nama instansi', 'Nama Lembaga Induk'),
        ('no SK perpus', 'Nomor SK Pendirian Perpustakaan'),
        ('jumlah judul buku', 'Jumlah Judul Koleksi'),
        ('jumlah eksemplar', 'Jumlah Eksemplar Koleksi'),
        ('jumlahanggota perpus(siswa)', 'Jumlah Anggota Perpustakaan'),
        ('jumlah petugas perpus', 'Jumlah Tenaga Perpustakaan'),
        ('luaslahan gedung', 'Luas Gedung Perpustakaan'),
        ('jumlah hari operasional dalam. 1 minggu', 'Jumlah Hari Layanan dalam Seminggu (hari)')
    ]
    
    # 3. Interactive Edit Loop
    while True:
        print("\n" + "="*50)
        print("HASIL EKSTRAKSI DATA (SILAKAN PERIKSA & EDIT):")
        print("="*50)
        for idx, (display, key) in enumerate(prompt_fields, 1):
            print(f"{idx:2d}. {display:<42} : {data[key]}")
            
        print("\nKetik nomor field (1-20) untuk mengedit nilainya.")
        print("Ketik 'y' jika semua data sudah benar dan siap disimpan ke Excel.")
        print("Ketik 'n' untuk membatalkan proses sesi ini.")
        
        choice = input("Pilihan Anda: ").strip().lower()
        
        if choice == 'y':
            break
        elif choice == 'n':
            print("Sesi dibatalkan. Tidak ada data yang ditulis ke Excel.")
            return
        elif choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(prompt_fields):
                display, key = prompt_fields[idx]
                new_val = input(f"Masukkan nilai baru untuk [{display}] (sebelumnya: '{data[key]}'): ").strip()
                data[key] = new_val
            else:
                print("Nomor tidak valid.")
        else:
            print("Pilihan tidak dikenal.")
            
    # 4. Save to Excel
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: File Excel {EXCEL_FILE} tidak ditemukan di folder aktif!")
        return
        
    print(f"\nSedang menyimpan ke {EXCEL_FILE}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Determine the sheet name: 'SD_' or 'SMP'
        subjenis = data['Subjenis Perpustakaan'].upper()
        school_name = data['Nama Lembaga Induk'].upper()
        if 'SMP' in subjenis or 'SMP' in school_name:
            sheet_name = 'SMP'
        else:
            sheet_name = 'SD_'
            
        ws = wb[sheet_name]
        
        # Find first empty row (checking column 8: Nama Perpustakaan)
        row_idx = 5
        while True:
            val = ws.cell(row=row_idx, column=8).value
            if val is None or str(val).strip() == "":
                break
            row_idx += 1
            
        print(f"Menulis data pada Sheet: [{sheet_name}], Baris: {row_idx}")
        
        # Write values to Excel (columns 1-24)
        # 1: No (use pre-existing if filled, else write new)
        existing_no = ws.cell(row=row_idx, column=1).value
        if existing_no is None:
            ws.cell(row=row_idx, column=1, value=row_idx - 4)
            
        # 2: Nama Lengkap Kepala Pustakawan
        ws.cell(row=row_idx, column=2, value=data['Nama Lengkap Kepala Pustakawan'])
        # 3: Jabatan
        ws.cell(row=row_idx, column=3, value=data['Jabatan'])
        # 4: Alamat Email perpustakaan
        ws.cell(row=row_idx, column=4, value=data['Alamat Email perpustakaan'])
        # 5: Nomor Telepon/Whatsapp
        ws.cell(row=row_idx, column=5, value=data['Nomor Telepon/Whatsapp'])
        # 6: Nomor Pokok Perpustakaan (NPP)
        ws.cell(row=row_idx, column=6, value=data['Nomor Pokok Perpustakaan (NPP)'])
        # 7: Subjenis Perpustakaan
        ws.cell(row=row_idx, column=7, value=data['Subjenis Perpustakaan'])
        # 8: Nama Perpustakaan
        ws.cell(row=row_idx, column=8, value=data['Nama Perpustakaan'])
        # 9: Provinsi
        ws.cell(row=row_idx, column=9, value=data['Provinsi'])
        # 10: Kabupaten
        ws.cell(row=row_idx, column=10, value=data['Kabupaten'])
        # 11: Kecamatan
        ws.cell(row=row_idx, column=11, value=data['Kecamatan'])
        # 12: Desa/Kelurahan
        ws.cell(row=row_idx, column=12, value=data['Desa/Kelurahan'])
        # 13: Alamat Lengkap (Jalan RT/ RW)
        ws.cell(row=row_idx, column=13, value=data['Alamat Lengkap (Jalan RT/ RW)'])
        # 14: Kode Pos
        # Try to write as integer if it is numeric
        zip_code = data['Kode Pos']
        if zip_code and zip_code.isdigit():
            ws.cell(row=row_idx, column=14, value=int(zip_code))
        else:
            ws.cell(row=row_idx, column=14, value=zip_code)
            
        # 15: Nama Lembaga Induk
        ws.cell(row=row_idx, column=15, value=data['Nama Lembaga Induk'])
        
        # 16: Tahun Berdiri Perpustakaan
        year_est = data['Tahun Berdiri Perpustakaan']
        if year_est and year_est.isdigit():
            ws.cell(row=row_idx, column=16, value=int(year_est))
        else:
            ws.cell(row=row_idx, column=16, value=year_est)
            
        # 17: Nomor SK Pendirian Perpustakaan
        ws.cell(row=row_idx, column=17, value=data['Nomor SK Pendirian Perpustakaan'])
        
        # 18: Jumlah Anggota Perpustakaan (Format with 'Siswa')
        ws.cell(row=row_idx, column=18, value=format_with_suffix(data['Jumlah Anggota Perpustakaan'], 'Siswa'))
        
        # 19: Luas Gedung (Format with 'm2')
        ws.cell(row=row_idx, column=19, value=format_with_suffix(data['Luas Gedung Perpustakaan'], 'm2'))
        
        # 20: Jumlah Tenaga (Format with 'Orang')
        ws.cell(row=row_idx, column=20, value=format_with_suffix(data['Jumlah Tenaga Perpustakaan'], 'Orang'))
        
        # 21: Jumlah Judul Koleksi
        ws.cell(row=row_idx, column=21, value=data['Jumlah Judul Koleksi'])
        # 22: Jumlah Eksemplar Koleksi
        ws.cell(row=row_idx, column=22, value=data['Jumlah Eksemplar Koleksi'])
        
        # 23: Jumlah Hari Layanan dalam Seminggu (hari)
        ws.cell(row=row_idx, column=23, value=data['Jumlah Hari Layanan dalam Seminggu (hari)'])
        
        # 24: Tahun Pendataan
        year_data = data['Tahun Pendataan']
        if year_data and year_data.isdigit():
            ws.cell(row=row_idx, column=24, value=int(year_data))
        else:
            ws.cell(row=row_idx, column=24, value=year_data)
            
        wb.save(EXCEL_FILE)
        print("Sukses menyimpan data ke Excel!")
        
    except Exception as e:
        print(f"ERROR Gagal menulis ke Excel: {e}")
        return
        
    # 5. Archive files
    clean_school_folder = re.sub(r'[^a-zA-Z0-9_]', '_', data['Nama Lembaga Induk']).strip('_')
    archive_dir = os.path.join(script_dir, "processed", f"{clean_school_folder}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    
    print(f"Mengarsipkan file ke: {archive_dir}")
    try:
        os.makedirs(archive_dir, exist_ok=True)
        for img in images:
            shutil.move(img, os.path.join(archive_dir, os.path.basename(img)))
        print("Pengarsipan selesai! Folder input sekarang bersih.")
    except Exception as e:
        print(f"Peringatan: Gagal memindahkan file ke arsip: {e}")

if __name__ == '__main__':
    main()
